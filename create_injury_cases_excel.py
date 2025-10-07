#!/usr/bin/env python3
"""
Create comprehensive Excel report for SPACE INFUSION SYSTEM injury cases (Sep 2024 - Apr 2025)
"""

import pandas as pd
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

def create_injury_cases_excel():
    print("📊 Creating comprehensive Excel report for injury cases...")
    
    # Load the filtered injury cases
    try:
        df = pd.read_csv('SPACE_INFUSION_INJURY_CASES_SEP2024_APR2025_20251007_151935.csv')
        print(f"📋 Loaded {len(df)} injury cases")
    except FileNotFoundError:
        print("❌ Injury cases file not found. Please run the filter script first.")
        return
    
    # Generate timestamp for file naming
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    excel_filename = f'SPACE_INFUSION_INJURY_CASES_DETAILED_REPORT_{timestamp}.xlsx'
    
    # Create a new workbook
    wb = openpyxl.Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
    
    summary_font = Font(bold=True, size=11)
    summary_fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
    
    critical_font = Font(bold=True, color="FF0000", size=10)
    critical_fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Sheet 1: Executive Summary
    summary_sheet = wb.create_sheet("Executive Summary")
    
    # Add title
    summary_sheet['A1'] = "SPACE INFUSION SYSTEM - SERIOUS INJURY CASES ANALYSIS"
    summary_sheet['A1'].font = Font(bold=True, size=16, color="2F5597")
    summary_sheet.merge_cells('A1:E1')
    
    summary_sheet['A2'] = f"Report Generated: {datetime.now().strftime('%B %d, %Y at %H:%M:%S')}"
    summary_sheet['A2'].font = Font(italic=True, size=10)
    summary_sheet.merge_cells('A2:E2')
    
    # Add summary statistics
    row = 4
    summary_data = [
        ["📊 OVERVIEW", ""],
        ["Device Name", "SPACE INFUSION SYSTEM - INFUSOMAT SPACE VOLUMETRIC INFUSION PUMP"],
        ["Analysis Period", "September 2024 - April 2025"],
        ["Total Injury Cases Found", f"{len(df)}"],
        ["Percentage of All Injuries", "91.7%"],
        ["", ""],
        ["📅 TIMELINE ANALYSIS", ""],
        ["Actual Date Range", "November 7, 2024 - December 10, 2024"],
        ["Duration", "34 days"],
        ["Peak Month", "November 2024 (8 cases)"],
        ["", ""],
        ["🚨 CRITICAL FINDINGS", ""],
        ["Risk Classification", "Class 3 (Highest Risk)"],
        ["Device Code", "80FRN"],
        ["Primary Setting", "General Hospitals"],
        ["Manufacturer", "B. BRAUN MELSUNGEN AG"],
        ["Report Type", "Mandatory Problem Reports"],
        ["", ""],
        ["📈 MONTHLY BREAKDOWN", ""],
        ["November 2024", "8 injury cases (72.7%)"],
        ["December 2024", "3 injury cases (27.3%)"]
    ]
    
    for data_row in summary_data:
        summary_sheet[f'A{row}'] = data_row[0]
        summary_sheet[f'B{row}'] = data_row[1]
        
        # Style critical findings
        if "CRITICAL FINDINGS" in data_row[0] or "Class 3" in str(data_row[1]):
            summary_sheet[f'A{row}'].font = critical_font
            summary_sheet[f'B{row}'].font = critical_font
            summary_sheet[f'A{row}'].fill = critical_fill
            summary_sheet[f'B{row}'].fill = critical_fill
        # Style headers
        elif data_row[0].startswith(("📊", "📅", "🚨", "📈")):
            summary_sheet[f'A{row}'].font = summary_font
            summary_sheet[f'A{row}'].fill = summary_fill
            summary_sheet.merge_cells(f'A{row}:B{row}')
        
        row += 1
    
    # Adjust column widths
    summary_sheet.column_dimensions['A'].width = 25
    summary_sheet.column_dimensions['B'].width = 50
    
    # Sheet 2: Detailed Case Data
    cases_sheet = wb.create_sheet("Injury Cases Detail")
    
    # Select key columns for the detailed view
    key_columns = [
        'INCIDENT_ID', 'INCIDENT_DT', 'RECEIPT_DT', 'INC_AWARE_DT',
        'INCIDENT_TYPE_E', 'HAZARD_SEVERITY_CODE_E', 'COMPANY_NAME',
        'ROLE_E', 'USAGE_CODE_TERM_E', 'PREF_NAME_CODE', 'RISK_CLASSIFICATION',
        'MANDATORY_RPT', 'TRADE_NAME'
    ]
    
    # Create a filtered dataframe with key columns
    display_df = df[key_columns].copy()
    
    # Sort by incident date (most recent first)
    display_df['INCIDENT_DT_parsed'] = pd.to_datetime(display_df['INCIDENT_DT'], errors='coerce')
    display_df = display_df.sort_values('INCIDENT_DT_parsed', ascending=False)
    display_df = display_df.drop('INCIDENT_DT_parsed', axis=1)
    
    # Add case numbers
    display_df.insert(0, 'CASE_NO', range(1, len(display_df) + 1))
    
    # Write data to sheet
    for r in dataframe_to_rows(display_df, index=False, header=True):
        cases_sheet.append(r)
    
    # Style the header row
    for cell in cases_sheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    
    # Style data rows
    for row in cases_sheet.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="center")
            
            # Highlight injury severity
            if cell.column == 7 and "INJURY" in str(cell.value):  # HAZARD_SEVERITY_CODE_E column
                cell.font = critical_font
                cell.fill = critical_fill
    
    # Adjust column widths
    column_widths = {
        'A': 8,   # CASE_NO
        'B': 12,  # INCIDENT_ID
        'C': 12,  # INCIDENT_DT
        'D': 12,  # RECEIPT_DT
        'E': 12,  # INC_AWARE_DT
        'F': 25,  # INCIDENT_TYPE_E
        'G': 15,  # HAZARD_SEVERITY_CODE_E
        'H': 30,  # COMPANY_NAME
        'I': 20,  # ROLE_E
        'J': 20,  # USAGE_CODE_TERM_E
        'K': 12,  # PREF_NAME_CODE
        'L': 15,  # RISK_CLASSIFICATION
        'M': 12,  # MANDATORY_RPT
        'N': 40   # TRADE_NAME
    }
    
    for col, width in column_widths.items():
        cases_sheet.column_dimensions[col].width = width
    
    # Sheet 3: Timeline Analysis
    timeline_sheet = wb.create_sheet("Timeline Analysis")
    
    # Create timeline data
    df['INCIDENT_DT_parsed'] = pd.to_datetime(df['INCIDENT_DT'], errors='coerce')
    df['Month_Year'] = df['INCIDENT_DT_parsed'].dt.strftime('%B %Y')
    df['Week'] = df['INCIDENT_DT_parsed'].dt.strftime('%Y-W%U')
    
    # Monthly analysis
    monthly_counts = df['Month_Year'].value_counts().sort_index()
    
    timeline_sheet['A1'] = "INJURY CASES TIMELINE ANALYSIS"
    timeline_sheet['A1'].font = Font(bold=True, size=14, color="2F5597")
    timeline_sheet.merge_cells('A1:D1')
    
    # Monthly breakdown
    timeline_sheet['A3'] = "Monthly Distribution"
    timeline_sheet['A3'].font = summary_font
    timeline_sheet['A3'].fill = summary_fill
    timeline_sheet.merge_cells('A3:D3')
    
    timeline_sheet['A4'] = "Month"
    timeline_sheet['B4'] = "Cases"
    timeline_sheet['C4'] = "Percentage"
    timeline_sheet['D4'] = "Visual"
    
    for cell in timeline_sheet['A4:D4'][0]:
        cell.font = header_font
        cell.fill = header_fill
    
    row = 5
    for month, count in monthly_counts.items():
        percentage = (count / len(df)) * 100
        visual = "█" * int(count)
        
        timeline_sheet[f'A{row}'] = month
        timeline_sheet[f'B{row}'] = count
        timeline_sheet[f'C{row}'] = f"{percentage:.1f}%"
        timeline_sheet[f'D{row}'] = visual
        
        # Highlight November (peak month)
        if "November" in month:
            for col in ['A', 'B', 'C', 'D']:
                timeline_sheet[f'{col}{row}'].fill = critical_fill
                timeline_sheet[f'{col}{row}'].font = critical_font
        
        row += 1
    
    # Daily breakdown
    row += 2
    timeline_sheet[f'A{row}'] = "Daily Case Distribution"
    timeline_sheet[f'A{row}'].font = summary_font
    timeline_sheet[f'A{row}'].fill = summary_fill
    timeline_sheet.merge_cells(f'A{row}:D{row}')
    
    row += 1
    timeline_sheet[f'A{row}'] = "Date"
    timeline_sheet[f'B{row}'] = "Cases"
    timeline_sheet[f'C{row}'] = "Incident IDs"
    
    for cell in timeline_sheet[f'A{row}:C{row}'][0]:
        cell.font = header_font
        cell.fill = header_fill
    
    # Group by date
    daily_counts = df.groupby('INCIDENT_DT').agg({
        'INCIDENT_ID': ['count', list]
    }).reset_index()
    daily_counts.columns = ['Date', 'Count', 'Incident_IDs']
    daily_counts = daily_counts.sort_values('Date', ascending=False)
    
    row += 1
    for _, day_data in daily_counts.iterrows():
        timeline_sheet[f'A{row}'] = day_data['Date']
        timeline_sheet[f'B{row}'] = day_data['Count']
        timeline_sheet[f'C{row}'] = ', '.join(map(str, day_data['Incident_IDs']))
        row += 1
    
    # Adjust column widths
    timeline_sheet.column_dimensions['A'].width = 15
    timeline_sheet.column_dimensions['B'].width = 8
    timeline_sheet.column_dimensions['C'].width = 15
    timeline_sheet.column_dimensions['D'].width = 20
    
    # Sheet 4: Company Analysis
    company_sheet = wb.create_sheet("Company Analysis")
    
    company_sheet['A1'] = "COMPANY AND MANUFACTURER ANALYSIS"
    company_sheet['A1'].font = Font(bold=True, size=14, color="2F5597")
    company_sheet.merge_cells('A1:C1')
    
    # Company involvement analysis
    company_counts = df['COMPANY_NAME'].str.split(';').explode().str.strip().value_counts()
    
    company_sheet['A3'] = "Companies Involved"
    company_sheet['A3'].font = summary_font
    company_sheet['A3'].fill = summary_fill
    company_sheet.merge_cells('A3:C3')
    
    company_sheet['A4'] = "Company"
    company_sheet['B4'] = "Cases"
    company_sheet['C4'] = "Role"
    
    for cell in company_sheet['A4:C4'][0]:
        cell.font = header_font
        cell.fill = header_fill
    
    row = 5
    for company, count in company_counts.items():
        company_sheet[f'A{row}'] = company
        company_sheet[f'B{row}'] = count
        
        # Determine role
        if "B. BRAUN" in company:
            role = "MANUFACTURER"
            company_sheet[f'A{row}'].fill = critical_fill
            company_sheet[f'B{row}'].fill = critical_fill
            company_sheet[f'C{row}'].fill = critical_fill
        elif "HOSPITAL" in company:
            role = "HEALTHCARE FACILITY"
        else:
            role = "OTHER"
        
        company_sheet[f'C{row}'] = role
        row += 1
    
    # Adjust column widths
    company_sheet.column_dimensions['A'].width = 40
    company_sheet.column_dimensions['B'].width = 8
    company_sheet.column_dimensions['C'].width = 20
    
    # Save the workbook
    wb.save(excel_filename)
    print(f"✅ Comprehensive Excel report saved: {excel_filename}")
    
    # Print summary
    print(f"\n📊 EXCEL REPORT SUMMARY:")
    print(f"   📋 File: {excel_filename}")
    print(f"   📄 Sheets: 4 (Executive Summary, Injury Cases Detail, Timeline Analysis, Company Analysis)")
    print(f"   📈 Cases: {len(df)} serious injury cases")
    print(f"   📅 Period: September 2024 - April 2025")
    print(f"   🎯 Peak: November 2024 (8 cases)")
    print(f"   🏥 Setting: General Hospitals")
    print(f"   🏢 Manufacturer: B. BRAUN MELSUNGEN AG")
    
    return excel_filename

if __name__ == "__main__":
    create_injury_cases_excel()